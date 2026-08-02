"""Import EBL deposit and lending interest rates from one Excel workbook.

Usage:
    python backend\import_rates_excel.py EBL_rates_update_template.xlsx
"""

from pathlib import Path
import argparse
from numbers import Number
import sqlite3
import sys

from openpyxl import load_workbook

from deposit_rate_database import (
    DATABASE_PATH,
    DEPOSIT_RATE_COLUMNS,
    build_search_text as build_deposit_search_text,
    create_deposit_rate_table,
)
from lending_rate_database import (
    LENDING_RATE_COLUMNS,
    build_search_text as build_lending_search_text,
    create_lending_rate_table,
)


RATE_DATA_DIR = Path(__file__).resolve().parent / "rate_data"
DEFAULT_RATE_EXCEL_PATH = RATE_DATA_DIR / "EBL_rates_update_template.xlsx"
DEPOSIT_SHEET_NAME = "Deposit Rates"
LENDING_SHEET_NAME = "Lending Rates"

DEPOSIT_REQUIRED_COLUMNS = [
    "business_unit",
    "category",
    "product",
    "condition",
    "rate",
    "source_file",
]
LENDING_REQUIRED_COLUMNS = [
    "section",
    "economic_purpose",
    "declared_rate",
    "lowest_rate",
    "highest_rate",
    "source_file",
]


def clean_cell(value):
    if value is None:
        return ""

    return str(value).strip()


def clean_excel_cell(cell, column):
    value = cell.value

    if value is None:
        return ""

    rate_columns = {
        "rate",
        "declared_rate",
        "lowest_rate",
        "highest_rate",
    }

    if (
        column in rate_columns
        and isinstance(value, Number)
        and "%" in (cell.number_format or "")
    ):
        return f"{value * 100:.2f}%"

    return clean_cell(value)


def resolve_excel_path(raw_path):
    path = Path(raw_path)

    if not path.is_absolute():
        path = Path.cwd() / path

    if path.exists():
        return path.resolve()

    fallback_path = RATE_DATA_DIR / Path(raw_path).name

    if fallback_path.exists():
        return fallback_path.resolve()

    return path.resolve()


def header_map(header_row, columns, sheet_name):
    headers = {}

    for column_index, value in enumerate(header_row, start=1):
        header = clean_cell(value).lower()

        if header:
            headers[header] = column_index

    missing_columns = [
        column
        for column in columns
        if column not in headers
    ]

    if missing_columns:
        raise ValueError(
            f"{sheet_name} sheet missing columns: "
            + ", ".join(missing_columns)
        )

    return headers


def read_rate_sheet(sheet, columns, required_columns, sheet_name):
    row_iterator = sheet.iter_rows()
    header_row = next(row_iterator, None)

    if not header_row:
        raise ValueError(f"{sheet_name} sheet is empty")

    headers = header_map(
        [cell.value for cell in header_row],
        columns,
        sheet_name,
    )
    rows = []

    for row_number, cells in enumerate(row_iterator, start=2):
        row = {
            column: clean_excel_cell(cells[headers[column] - 1], column)
            if len(cells) >= headers[column]
            else ""
            for column in columns
        }

        if not any(row.values()):
            continue

        missing_required = [
            column
            for column in required_columns
            if not row[column]
        ]

        if missing_required:
            raise ValueError(
                f"{sheet_name} row {row_number} missing required values: "
                + ", ".join(missing_required)
            )

        rows.append(row)

    if not rows:
        raise ValueError(f"{sheet_name} sheet has no rows to import")

    return rows


def read_rates_excel(path):
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        for sheet_name in [DEPOSIT_SHEET_NAME, LENDING_SHEET_NAME]:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Workbook does not have a '{sheet_name}' sheet")

        deposit_rows = read_rate_sheet(
            workbook[DEPOSIT_SHEET_NAME],
            DEPOSIT_RATE_COLUMNS,
            DEPOSIT_REQUIRED_COLUMNS,
            DEPOSIT_SHEET_NAME,
        )
        lending_rows = read_rate_sheet(
            workbook[LENDING_SHEET_NAME],
            LENDING_RATE_COLUMNS,
            LENDING_REQUIRED_COLUMNS,
            LENDING_SHEET_NAME,
        )

        return deposit_rows, lending_rows

    finally:
        workbook.close()


def import_rate_rows(deposit_rows, lending_rows):
    connection = sqlite3.connect(DATABASE_PATH)
    create_deposit_rate_table(connection)
    create_lending_rate_table(connection)
    cursor = connection.cursor()

    try:
        cursor.execute("DELETE FROM deposit_rates")
        cursor.execute("DELETE FROM lending_rates")

        for row in deposit_rows:
            cursor.execute("""
                INSERT INTO deposit_rates (
                    business_unit,
                    category,
                    product,
                    condition,
                    rate,
                    note,
                    source_file,
                    search_text
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row["business_unit"],
                row["category"],
                row["product"],
                row["condition"],
                row["rate"],
                row["note"],
                row["source_file"],
                build_deposit_search_text(row),
            ))

        for row in lending_rows:
            cursor.execute("""
                INSERT INTO lending_rates (
                    section,
                    category,
                    subcategory,
                    economic_purpose,
                    declared_rate,
                    lowest_rate,
                    highest_rate,
                    pdf_page,
                    source_file,
                    search_text
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row["section"],
                row["category"],
                row["subcategory"],
                row["economic_purpose"],
                row["declared_rate"],
                row["lowest_rate"],
                row["highest_rate"],
                row["pdf_page"],
                row["source_file"],
                build_lending_search_text(row),
            ))

        connection.commit()

    except Exception:
        connection.rollback()
        raise

    finally:
        connection.close()

    return {
        "deposit": len(deposit_rows),
        "lending": len(lending_rows),
    }


def import_rates_excel(path):
    deposit_rows, lending_rows = read_rates_excel(path)
    return import_rate_rows(deposit_rows, lending_rows)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import EBL deposit and lending interest rates from Excel."
    )
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=str(DEFAULT_RATE_EXCEL_PATH),
        help=(
            "Path to the Excel workbook. Defaults to "
            "backend/rate_data/EBL_rates_update_template.xlsx"
        ),
    )

    return parser.parse_args()


def main():
    args = parse_args()
    excel_path = resolve_excel_path(args.excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    inserted = import_rates_excel(excel_path)
    print(
        "Imported "
        f"{inserted['deposit']} deposit rate rows and "
        f"{inserted['lending']} lending rate rows into {DATABASE_PATH}"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Import failed: {error}", file=sys.stderr)
        raise SystemExit(1)

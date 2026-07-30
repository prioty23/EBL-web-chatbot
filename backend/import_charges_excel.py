"""Import Schedule of Charges from an Excel workbook into SQLite.

Usage:
    python backend\import_charges_excel.py EBL_charges_update_template.xlsx
"""

from pathlib import Path
import argparse
import sqlite3
import sys

from openpyxl import load_workbook

from charge_database import (
    CHARGE_COLUMNS,
    DATABASE_PATH,
    build_search_text,
    create_charge_table,
)


REQUIRED_COLUMNS = [
    "schedule",
    "category",
    "product",
    "charge_name",
    "amount",
    "source_file",
]
ALLOWED_SCHEDULES = {"Retail", "SME", "Corporate", "Cards"}


def clean_cell(value):
    if value is None:
        return ""

    return str(value).strip()


def resolve_excel_path(raw_path):
    path = Path(raw_path)

    if not path.is_absolute():
        path = Path.cwd() / path

    if path.exists():
        return path.resolve()

    fallback_path = Path(__file__).resolve().parent / "charge_data" / Path(raw_path).name

    if fallback_path.exists():
        return fallback_path.resolve()

    return path.resolve()


def header_map(header_row):
    headers = {}

    for column_index, value in enumerate(header_row, start=1):
        header = clean_cell(value).lower()

        if header:
            headers[header] = column_index

    missing_columns = [
        column
        for column in CHARGE_COLUMNS
        if column not in headers
    ]

    if missing_columns:
        raise ValueError(
            "Charges sheet missing columns: "
            + ", ".join(missing_columns)
        )

    return headers


def read_charge_excel(path, sheet_name="Charges"):
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook does not have a '{sheet_name}' sheet")

        sheet = workbook[sheet_name]
        row_iterator = sheet.iter_rows(values_only=True)
        header_row = next(row_iterator, None)

        if not header_row:
            raise ValueError("Charges sheet is empty")

        headers = header_map(header_row)
        rows = []

        for row_number, values in enumerate(row_iterator, start=2):
            row = {
                column: clean_cell(values[headers[column] - 1])
                if len(values) >= headers[column]
                else ""
                for column in CHARGE_COLUMNS
            }

            if not any(row.values()):
                continue

            missing_required = [
                column
                for column in REQUIRED_COLUMNS
                if not row[column]
            ]

            if missing_required:
                raise ValueError(
                    f"Charges row {row_number} missing required values: "
                    + ", ".join(missing_required)
                )

            if row["schedule"] not in ALLOWED_SCHEDULES:
                raise ValueError(
                    f"Charges row {row_number} has invalid schedule "
                    f"'{row['schedule']}'. Use Retail, SME, Corporate, or Cards."
                )

            rows.append(row)

        if not rows:
            raise ValueError("Charges sheet has no charge rows to import")

        return rows

    finally:
        workbook.close()


def import_charge_rows(rows):
    connection = sqlite3.connect(DATABASE_PATH)
    create_charge_table(connection)
    cursor = connection.cursor()
    cursor.execute("DELETE FROM charges")

    for row in rows:
        cursor.execute("""
            INSERT INTO charges (
                schedule,
                category,
                product,
                charge_name,
                condition,
                amount,
                vat_note,
                source_file,
                search_text
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row["schedule"],
            row["category"],
            row["product"],
            row["charge_name"],
            row["condition"],
            row["amount"],
            row["vat_note"],
            row["source_file"],
            build_search_text(row),
        ))

    connection.commit()
    connection.close()

    return len(rows)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import EBL schedule of charges from Excel into SQLite."
    )
    parser.add_argument(
        "excel_path",
        help="Path to the Excel workbook, for example EBL_charges_update_template.xlsx",
    )

    return parser.parse_args()


def main():
    args = parse_args()
    excel_path = resolve_excel_path(args.excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    rows = read_charge_excel(excel_path)
    inserted = import_charge_rows(rows)

    print(f"Imported {inserted} charge rows into {DATABASE_PATH}")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Import failed: {error}", file=sys.stderr)
        raise SystemExit(1)
